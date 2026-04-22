"""Text inspector tool for reading attached files.

This tool allows the model to read text content from attached files,
supporting multiple file formats including Office documents, PDFs, and data files.

Supports two modes:
1. Direct mode: Returns raw file content
2. Sub-agent mode: Uses LLM to analyze file and answer questions (with thinking tag stripping)
"""

import html
import os
import re
import string
import zipfile
from pathlib import Path
from typing import Any, Dict, Optional

from ..core.tool import BaseTool, ToolResult
from ..utils.logging import get_logger
from ..utils.parsing import strip_thinking_tags

logger = get_logger(__name__)

# Supported file extensions
SUPPORTED_TEXT_EXTS = {
    ".txt", ".md", ".log",
    ".json", ".jsonl",
    ".xml",
    ".csv", ".tsv",
    ".yaml", ".yml",
    ".docx",
    ".xlsx",
    ".jsonld",
    ".parquet",
    ".pdf",
    ".pdb",
    ".pptx",
    ".py",
}


class TextInspectorTool(BaseTool):
    """Tool for inspecting text files.

    Reads text content from files attached to questions.
    Supports multiple file formats including Office documents, PDFs, and data files.

    Two modes:
    - Direct mode (model_provider=None): Returns raw file content only
    - Sub-agent mode (model_provider set): Uses LLM to analyze file and answer questions
    """

    def __init__(
        self,
        max_chars: int = 50000,
        model_provider: Optional[Any] = None,
        use_thinking: bool = False
    ):
        """Initialize text inspector tool.

        Args:
            max_chars: Maximum characters to read from file
            model_provider: Optional model provider for sub-agent analysis
            use_thinking: Whether to use thinking mode for LLM
        """
        self.max_chars = max_chars
        self.model_provider = model_provider
        self.use_thinking = use_thinking
        self.direct_mode = model_provider is None

    @property
    def name(self) -> str:
        return "text_inspector"

    @property
    def description(self) -> str:
        if self.direct_mode:
            return "Read text content from attached files"
        else:
            return "Read and analyze text content from attached files with optional question answering"

    def get_schema(self) -> Dict[str, Any]:
        """Return Qwen3 JSON Schema.

        Schema differs by mode:
        - Direct mode: No parameters (framework injects full_file_path; returns raw content)
        - Non-direct mode: Optional question parameter only (framework injects full_file_path)
        """
        if self.direct_mode:
            # Direct mode: no parameters - return raw file content only
            return {
                "type": "function",
                "function": {
                    "name": "text_inspector",
                    "description": (
                        "Return the raw text of the attached file for this question. "
                        "Use this tool only to retrieve the file contents, and reason about them yourself."
                    ),
                    "parameters": {
                        "type": "object",
                        "properties": {},
                        "required": []
                    }
                }
            }
        else:
            # Non-direct mode: optional question for LLM analysis
            return {
                "type": "function",
                "function": {
                    "name": "text_inspector",
                    "description": (
                        "Inspect the attached plain-text file for this question. "
                        "Important: Do NOT guess or provide file paths. The system will automatically use the "
                        "question's attached file. Optionally ask a question about the file. If you don't have questions about the file and you only want to see the file text, don't include the question parameter."
                    ),
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "question": {
                                "type": "string",
                                "description": "Optional question about the attached file. If omitted, return the file text."
                            }
                        },
                        "required": []
                    }
                }
            }

    def execute(
        self,
        full_file_path: str,
        question: Optional[str] = None,
        shared_context: str = "",
    ) -> ToolResult:
        """Read text from file.

        Args:
            full_file_path: Path to text file (injected by orchestrator from attachments)
            question: Optional question about the file (non-direct mode only)
            shared_context: Optional orchestrator shared-memory block passed
                through to the sub-agent analysis prompt (ignored in direct
                mode or when no question is provided). Empty string disables
                the feature.

        Returns:
            ToolResult with file content or LLM analysis
        """
        logger.info(f"Inspecting text file: {full_file_path}")

        path = Path(full_file_path)

        # Check if file exists
        if not path.exists():
            msg = f"[text_inspector] File not found: {full_file_path}"
            return ToolResult(
                success=False,
                output=msg,
                metadata={"file_path": full_file_path},
                error=msg,
            )

        # Check if it's a file (not directory)
        if not path.is_file():
            msg = f"[text_inspector] Path is not a file: {full_file_path}"
            return ToolResult(
                success=False,
                output=msg,
                metadata={"file_path": full_file_path},
                error=msg,
            )

        # Get file extension
        file_ext = path.suffix.lower()

        # Check if supported
        if file_ext not in SUPPORTED_TEXT_EXTS:
            msg = (
                f"[text_inspector] Unsupported file type: {file_ext}. "
                f"Supported types: {', '.join(sorted(SUPPORTED_TEXT_EXTS))}"
            )
            return ToolResult(
                success=False,
                output=msg,
                metadata={"file_path": full_file_path, "file_type": file_ext},
                error=msg,
            )

        # Read file content based on type
        try:
            if file_ext == ".docx":
                content = self._read_docx(path)
            elif file_ext == ".xlsx":
                content = self._read_xlsx(path)
            elif file_ext == ".pdf":
                content = self._read_pdf(path)
            elif file_ext == ".pptx":
                content = self._read_pptx(path)
            elif file_ext == ".parquet":
                content = self._read_parquet(path)
            else:
                # Plain text file
                content = self._read_plain_text(path)

            # Truncate if necessary
            truncated = False
            if len(content) > self.max_chars:
                content = content[:self.max_chars]
                truncated = True

            # Get file info
            file_size = path.stat().st_size

            # Format output based on mode
            if self.direct_mode or not question:
                # Direct mode or no question: return raw file content
                output = f"File: {path.name}\n"
                output += f"Size: {file_size} bytes\n"
                output += f"Type: {file_ext}\n"
                output += f"\n{'='*60}\n"
                output += f"Content:\n{'='*60}\n\n"
                output += content

                if truncated:
                    output += f"\n\n[... truncated, showing first {self.max_chars} characters]"

                if not self.direct_mode and not question:
                    output += "\n\n[Note] No question provided; returned raw file content."

                return ToolResult(
                    success=True,
                    output=output,
                    metadata={
                        "file_path": full_file_path,
                        "file_name": path.name,
                        "file_size": file_size,
                        "file_type": file_ext,
                        "truncated": truncated,
                        "mode": "direct" if self.direct_mode else "raw"
                    }
                )
            else:
                # Non-direct mode with question: use LLM analysis
                analysis, usage = self._analyze_with_llm(content, question, shared_context=shared_context)

                return ToolResult(
                    success=True,
                    output=analysis,
                    metadata={
                        "file_path": full_file_path,
                        "file_name": path.name,
                        "file_size": file_size,
                        "file_type": file_ext,
                        "truncated": truncated,
                        "mode": "sub-agent",
                        "question": question
                    },
                    usage=usage,
                )

        except Exception as e:
            logger.error(f"Error reading file: {e}", exc_info=True)
            msg = f"[text_inspector] Error reading file {full_file_path}: {str(e)}"
            return ToolResult(
                success=False,
                output=msg,
                metadata={"file_path": full_file_path, "file_type": file_ext},
                error=msg,
            )

    def _analyze_with_llm(
        self,
        file_content: str,
        question: str,
        shared_context: str = "",
    ) -> tuple[str, Optional[Dict[str, int]]]:
        """Use LLM to analyze file content and answer question (sub-agent mode).

        Args:
            file_content: Content of the file
            question: Question about the file
            shared_context: Optional orchestrator shared-memory block prepended
                to the user prompt.

        Returns:
            Tuple of (analysis text, token usage dict or None)
        """
        if not self.model_provider:
            return "[Note] No text-inspector model configured; cannot analyze the file.", None

        system_prompt = (
            "You are given the content of a plain-text file attached to the user's question. "
            "Answer the question using only the file content. If the file does not contain the answer, say so."
        )
        shared_block = (
            "Shared context:\n\n"
            f"{shared_context}\n\n---\n\n"
        ) if shared_context else ""
        user_prompt = f"{shared_block}File content:\n\n{file_content}\n\nQuestion:\n{question}\n"

        prompt_messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ]

        prompt = self.model_provider.apply_chat_template(prompt_messages, use_thinking=self.use_thinking)
        result = self.model_provider.generate([prompt])[0]

        output = strip_thinking_tags(result.text)
        return output, result.usage

    def _read_plain_text(self, path: Path) -> str:
        """Read plain text file."""
        with open(path, 'r', encoding='utf-8', errors='replace') as f:
            return f.read()

    def _read_docx(self, path: Path) -> str:
        """Read DOCX file using zipfile (stdlib only)."""
        try:
            with zipfile.ZipFile(path, "r") as zf:
                with zf.open("word/document.xml") as f:
                    xml_bytes = f.read()
        except KeyError:
            raise ValueError("DOCX is missing word/document.xml")
        except zipfile.BadZipFile as exc:
            raise ValueError(f"Invalid DOCX (bad zip): {exc}")

        # Parse XML to extract text
        xml_text = xml_bytes.decode("utf-8", errors="replace")

        # Convert paragraph and break tags to newlines
        xml_text = re.sub(r"</w:p[^>]*>", "\n", xml_text, flags=re.IGNORECASE)
        xml_text = re.sub(r"<w:br[^>]*/>", "\n", xml_text, flags=re.IGNORECASE)
        xml_text = re.sub(r"<w:tab[^>]*/>", "\t", xml_text, flags=re.IGNORECASE)

        # Remove XML tags
        xml_text = re.sub(r"<[^>]+>", "", xml_text)
        text = html.unescape(xml_text)

        # Normalize line breaks
        text = re.sub(r"\r\n?", "\n", text)
        text = re.sub(r"\n{3,}", "\n\n", text).strip()

        return text

    def _read_pdf(self, path: Path, max_pages: int = 30) -> str:
        """Read PDF file."""
        try:
            from pypdf import PdfReader
        except ImportError:
            raise ImportError("pypdf is required for PDF support. Install with: pip install pypdf")

        reader = PdfReader(path)
        parts = []

        for i, page in enumerate(reader.pages):
            if max_pages and i >= max_pages:
                parts.append("[TRUNCATED_PAGES]")
                break

            try:
                txt = page.extract_text() or ""
            except Exception:
                txt = ""

            if txt:
                parts.append(txt)

        return "\n\n".join(parts).strip()

    def _read_pptx(self, path: Path, max_slides: int = 50) -> str:
        """Read PPTX file."""
        try:
            from pptx import Presentation
        except ImportError:
            raise ImportError("python-pptx is required for PPTX support. Install with: pip install python-pptx")

        prs = Presentation(path)
        parts = []

        for i, slide in enumerate(prs.slides):
            if max_slides and i >= max_slides:
                parts.append("[TRUNCATED_SLIDES]")
                break

            parts.append(f"[SLIDE {i+1}]")

            for shape in getattr(slide, "shapes", []):
                text = getattr(shape, "text", None)
                if text:
                    parts.append(text)

            parts.append("")

        return "\n".join(parts).strip()

    def _read_parquet(self, path: Path, max_rows: int = 200, max_cols: int = 50) -> str:
        """Read Parquet file."""
        try:
            import pandas as pd
        except ImportError:
            raise ImportError("pandas is required for Parquet support. Install with: pip install pandas pyarrow")

        df = pd.read_parquet(path)

        # Limit columns
        if max_cols and df.shape[1] > max_cols:
            df = df.iloc[:, :max_cols]

        # Limit rows
        if max_rows:
            df = df.head(max_rows)

        # Escape control characters
        def esc(x):
            s = "" if x is None else str(x)
            s = s.replace("\\", "\\\\")
            s = s.replace("\t", "\\t").replace("\r", "\\r").replace("\n", "\\n")
            return s

        # Format as TSV
        cols = [esc(c) for c in df.columns.tolist()]
        lines = ["\t".join(cols)]

        for row in df.itertuples(index=False, name=None):
            lines.append("\t".join(esc(v) for v in row))

        if max_rows and len(df) == max_rows:
            lines.append("[TRUNCATED_ROWS]")

        return "\n".join(lines).strip()

    def _read_xlsx(self, path: Path, max_rows: int = 200, max_cols: int = 50) -> str:
        """Read XLSX file using openpyxl, including both values and fill layout."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl is required for XLSX support. Install with: pip install openpyxl")

        file_path = str(path)
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        if not os.path.isfile(file_path):
            raise ValueError(f"Not a file: {file_path}")

        wb = load_workbook(filename=file_path, read_only=False, data_only=True)
        parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"[SHEET] {sheet_name}")

            # Determine used range (rows/cols), taking into account internal cell dict.
            used_max_r = int(getattr(ws, "max_row", 0) or 0)
            used_max_c = int(getattr(ws, "max_column", 0) or 0)
            try:
                cells_dict = getattr(ws, "_cells", None)
                if cells_dict:
                    used_max_r = max(used_max_r, max(r for (r, _c) in cells_dict.keys()))
                    used_max_c = max(used_max_c, max(c for (_r, c) in cells_dict.keys()))
            except Exception:
                pass

            if used_max_r <= 1 and used_max_c <= 1:
                max_r = max_rows if max_rows else 1
                max_c = max_cols if max_cols else 1
            else:
                max_r = min(used_max_r, max_rows) if max_rows else used_max_r
                max_c = min(used_max_c, max_cols) if max_cols else used_max_c

            # Palette + legend for fills (single-letter tokens → RGB/theme/index).
            _palette = {}
            _legend = {}
            _token_iter = iter(string.ascii_uppercase)

            def _safe_getattr(obj, name: str):
                try:
                    return getattr(obj, name, None)
                except Exception:
                    return None

            def _safe_str(x):
                if x is None:
                    return None
                try:
                    s = str(x)
                except Exception:
                    return None
                s = s.strip()
                if "Values must be of type" in s:
                    return None
                return s or None

            def _fill_key_and_desc(cell):
                fill = getattr(cell, "fill", None)
                if not fill:
                    return None, None
                pattern = _safe_getattr(fill, "patternType")

                fg = _safe_getattr(fill, "fgColor")
                rgb = _safe_getattr(fg, "rgb") if fg is not None else None
                index = _safe_getattr(fg, "index") if fg is not None else None
                theme = _safe_getattr(fg, "theme") if fg is not None else None
                tint = _safe_getattr(fg, "tint") if fg is not None else None
                typ = _safe_getattr(fg, "type") if fg is not None else None

                pattern_is_none = (pattern is None) or (str(pattern).lower() == "none")
                rgb_str = _safe_str(rgb)
                rgb_meaningful = rgb_str is not None and rgb_str != "00000000"
                has_any_color = rgb_meaningful or any(x is not None for x in (index, theme, tint, typ))
                if pattern_is_none and not has_any_color:
                    return None, None

                key = (
                    f"patternType={pattern}|fg.type={typ}|fg.rgb={rgb}|"
                    f"fg.index={index}|fg.theme={theme}|fg.tint={tint}"
                )
                if rgb_meaningful:
                    desc = rgb_str
                else:
                    theme_str = _safe_str(theme)
                    index_str = _safe_str(index)
                    if theme_str is not None:
                        desc = f"THEME:{theme_str}"
                    elif index_str is not None:
                        desc = f"INDEX:{index_str}"
                    else:
                        desc = "NO_RGB"
                return key, desc

            def _fill_to_token(cell):
                key, desc = _fill_key_and_desc(cell)
                if key is None:
                    return "."
                token = _palette.get(key)
                if token is None:
                    try:
                        token = next(_token_iter)
                    except StopIteration:
                        token = str(len(_palette) + 1)
                    _palette[key] = token
                    _legend[token] = desc
                return token

            has_values = False
            has_fills = False
            values_lines = []
            grid_lines = []

            for r in range(1, max_r + 1):
                row_tokens = []
                for c in range(1, max_c + 1):
                    cell = ws.cell(row=r, column=c)
                    tok = _fill_to_token(cell)
                    if tok != ".":
                        has_fills = True
                    row_tokens.append(tok)

                    v = cell.value
                    if v not in (None, ""):
                        has_values = True
                        s = str(v)
                        s = s.replace("\\", "\\\\")
                        s = s.replace("\t", "\\t").replace("\r", "\\r").replace("\n", "\\n")
                        values_lines.append(f"R{r}C{c}: {s}")
                grid_lines.append("".join(row_tokens).rstrip())

            if has_values:
                parts.append("[VALUES] (with coordinates)")
                parts.extend(values_lines)

            if has_fills:
                if has_values:
                    parts.append("")
                parts.append("[FILL_GRID] (tokens)")
                parts.append("[LEGEND] .=no fill; other tokens map to RGB in [FILL_LEGEND]")
                parts.extend(grid_lines)
                if _legend:
                    parts.append("")
                    parts.append("[FILL_LEGEND]")

                    def _legend_sort_key(tok):
                        return (0, tok) if tok.isalpha() else (1, int(tok)) if tok.isdigit() else (2, tok)

                    for tok in sorted(_legend.keys(), key=_legend_sort_key):
                        parts.append(f"[{tok}] {_legend[tok]}")

            if not has_values and not has_fills:
                parts.append("[Empty sheet]")

            if ws.max_row and max_rows and ws.max_row > max_rows:
                parts.append("[TRUNCATED_ROWS]")
            if ws.max_column and max_cols and ws.max_column > max_cols:
                parts.append("[TRUNCATED_COLS]")

            parts.append("")

        return "\n".join(parts).strip()

    def validate_args(self, **kwargs) -> bool:
        """Validate arguments.

        Args:
            **kwargs: Tool arguments

        Returns:
            True if valid
        """
        # Orchestrator injects `full_file_path` for attached files.
        file_path = kwargs.get("full_file_path")
        return isinstance(file_path, str) and len(file_path) > 0
